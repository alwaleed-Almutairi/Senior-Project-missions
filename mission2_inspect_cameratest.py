"""
mission2_inspect_cameratest.py
──────────────────────────────
Runs Mission 2 inspection using the REAL RealSense camera (color + depth)
but WITHOUT a drone connection. Drone movement is simulated.

This proves:
  - RealSense D435 color + depth streams work
  - Reads mission2_route.xlsx correctly
  - Captures close-up images at each crack location
  - Measures depth at each crack location
  - Generates final_report.xlsx with all data

Usage:
  python mission2_inspect_cameratest.py
"""

import asyncio
import cv2
import numpy as np
import pyrealsense2 as rs
import pandas as pd
import os
import time
from datetime import datetime


def load_cracks_from_excel(filename="mission2_route.xlsx"):
    cracks = []
    print(f"Reading crack coordinates from {filename} (Sheet: VisitOrder)...")
    try:
        df = pd.read_excel(filename, sheet_name="VisitOrder")
        for index, row in df.iterrows():
            if row["node_type"] != "CRACK":
                continue
            cracks.append({
                "id":         str(row["crack_id"]).strip(),
                "x":          float(row["x"]),
                "y":          float(row["y"]),
                "z":          float(row["z"]),
                "crack_type": str(row["crack_type"]) if "crack_type" in df.columns and pd.notna(row.get("crack_type")) else "Unknown",
                "confidence": float(row["confidence"]) if "confidence" in df.columns and pd.notna(row.get("confidence")) else 0.0,
            })
        print(f"Successfully loaded {len(cracks)} cracks.")
        return cracks
    except Exception as e:
        print(f"ERROR reading excel file: {e}")
        return []


async def stream_and_sleep(duration, pipeline):
    """Waits while keeping the RealSense color feed live."""
    end_time = asyncio.get_event_loop().time() + duration
    latencies = []

    while asyncio.get_event_loop().time() < end_time:
        t_start = time.perf_counter()
        frames = pipeline.poll_for_frames()
        if frames:
            color_frame = frames.get_color_frame()
            if color_frame:
                frame_data = np.asanyarray(color_frame.get_data())
                cv2.imshow("Ground Station - RealSense Live", frame_data)
                t_end = time.perf_counter()
                latencies.append((t_end - t_start) * 1000)
        cv2.waitKey(1)
        await asyncio.sleep(0.03)

    if latencies:
        avg = sum(latencies) / len(latencies)
        print(f"   [LATENCY] Avg: {avg:.1f} ms | Max: {max(latencies):.1f} ms | Frames: {len(latencies)}")


def generate_report(report_data, output_path="final_report.xlsx"):
    """Generates the final inspection report as a styled Excel file."""
    if not report_data:
        print("No inspection data to report.")
        return

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    df = pd.DataFrame(report_data)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Sheet 1: Full Report
        df.to_excel(writer, sheet_name="Report", index=False)
        ws = writer.sheets["Report"]

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(vertical="center")
                cell.border = thin_border

        col_widths = {
            "A": 12, "B": 10, "C": 10, "D": 10,
            "E": 20, "F": 12, "G": 8, "H": 8, "I": 8,
            "J": 12, "K": 35, "L": 20,
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Sheet 2: Summary
        summary_data = {
            "Metric": [
                "Total Cracks Inspected",
                "Inspection Date",
                "Crack Types Found",
                "Avg Confidence",
                "Avg Depth (mm)",
                "Min Depth (mm)",
                "Max Depth (mm)",
                "Images Saved",
            ],
            "Value": [
                len(df),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                ", ".join(sorted(df["crack_type"].unique())),
                f"{df['confidence'].mean():.4f}",
                f"{df['depth_mm'].mean():.2f}" if df["depth_mm"].sum() > 0 else "N/A",
                f"{df['depth_mm'].min():.2f}" if df["depth_mm"].sum() > 0 else "N/A",
                f"{df['depth_mm'].max():.2f}" if df["depth_mm"].sum() > 0 else "N/A",
                len(df[df["image_path"] != ""]),
            ],
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)
        ws2 = writer.sheets["Summary"]
        for col_idx in range(1, 3):
            cell = ws2.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 40

        # Sheet 3: By crack type
        type_summary = df.groupby("crack_type").agg(
            count=("crack_type", "size"),
            avg_confidence=("confidence", "mean"),
            avg_depth_mm=("depth_mm", "mean"),
        ).reset_index()
        type_summary.columns = ["Crack Type", "Count", "Avg Confidence", "Avg Depth (mm)"]
        type_summary.to_excel(writer, sheet_name="ByType", index=False)
        ws3 = writer.sheets["ByType"]
        for col_idx in range(1, 5):
            cell = ws3.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws3.column_dimensions["A"].width = 22
        ws3.column_dimensions["B"].width = 10
        ws3.column_dimensions["C"].width = 16
        ws3.column_dimensions["D"].width = 16

    print(f"\n{'='*50}")
    print(f"  FINAL REPORT: {output_path}")
    print(f"{'='*50}")
    print(f"  Cracks inspected : {len(df)}")
    print(f"  Crack types      : {', '.join(sorted(df['crack_type'].unique()))}")
    print(f"  Avg confidence   : {df['confidence'].mean():.2f}")
    if df["depth_mm"].sum() > 0:
        print(f"  Avg depth        : {df['depth_mm'].mean():.2f} mm")
    print(f"  Report saved to  : {os.path.abspath(output_path)}")
    print(f"{'='*50}")


async def run():
    print("=" * 60)
    print("  MISSION 2 — CAMERA TEST (No Drone Required)")
    print("  Tests: RealSense Color + Depth + Excel Pipeline")
    print("=" * 60)

    target_cracks = load_cracks_from_excel("mission2_route.xlsx")
    if not target_cracks:
        print("No cracks to inspect. Run mission2_solverTSP.py first.")
        return

    print("\nStarting Intel RealSense pipeline (Color + Depth)...")
    pipeline = rs.pipeline()
    config = rs.config()
    try:
        config.enable_stream(rs.stream.color, 640, 480, rs.format.bgr8, 30)
        config.enable_stream(rs.stream.depth, 640, 480, rs.format.z16, 30)
        pipeline.start(config)

        align_to = rs.stream.color
        align = rs.align(align_to)

        print("RealSense D435 initialized (Color + Depth)!")
        await stream_and_sleep(2, pipeline)
    except Exception as e:
        print(f"RealSense failed to initialize: {e}")
        return

    output_folder = "img_inspection"
    os.makedirs(output_folder, exist_ok=True)

    print("\n-- [SIM] Arming (simulated)")
    print("-- [SIM] Taking off (simulated)")
    await stream_and_sleep(3, pipeline)

    print("\n-- Starting Close-up Inspection")
    print("-" * 60)

    CLOSE_UP_OFFSET = 0.5
    report_data = []

    for i, crack in enumerate(target_cracks, 1):
        cid = crack["id"]
        n = crack["x"] + CLOSE_UP_OFFSET
        e = crack["y"]
        d = crack["z"]

        print(f"\n  [{i}/{len(target_cracks)}] Flying to Crack {cid} at [N:{n:.2f}, E:{e:.2f}, D:{d:.2f}]...")
        print(f"        Type: {crack['crack_type']} | Confidence: {crack['confidence']:.2f}")
        print("        [SIM] Drone moving to position...")
        await stream_and_sleep(4, pipeline)

        print(f"        Capturing color + depth of Crack {cid}...")
        frames = pipeline.wait_for_frames()
        aligned_frames = align.process(frames)
        color_frame = aligned_frames.get_color_frame()
        depth_frame = aligned_frames.get_depth_frame()

        depth_mm = 0.0
        filepath = ""

        if color_frame and depth_frame:
            color_image = np.asanyarray(color_frame.get_data())
            filepath = os.path.join(output_folder, f"crack_{cid}_closeup.jpg")
            cv2.imwrite(filepath, color_image)

            center_distance = depth_frame.get_distance(320, 240)
            depth_mm = round(center_distance * 1000, 2)

            print(f"        [DATA] Depth at center: {depth_mm:.2f} mm ({center_distance:.3f} m)")
            print(f"        [SAVE] {filepath}")
        else:
            print("        WARNING: RealSense dropped a frame during capture!")

        report_data.append({
            "waypoint_id":  cid,
            "north":        round(n, 6),
            "east":         round(e, 6),
            "down":         round(d, 6),
            "crack_type":   crack["crack_type"],
            "confidence":   round(crack["confidence"], 4),
            "X":            0.0,
            "Y":            round(crack["y"], 6),
            "Z":            round(crack["z"], 6),
            "depth_mm":     depth_mm,
            "image_path":   filepath,
            "timestamp":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    # Generate final report
    print("\n" + "-" * 60)
    generate_report(report_data, "final_report.xlsx")

    print("\n-- [SIM] Returning to home (simulated)")
    await stream_and_sleep(3, pipeline)
    print("-- [SIM] Landing (simulated)")

    pipeline.stop()
    cv2.destroyAllWindows()
    print("-- Done. RealSense stopped safely.")


if __name__ == "__main__":
    asyncio.run(run())
