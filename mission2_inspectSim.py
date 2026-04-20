import asyncio
import cv2
import numpy as np
import pandas as pd
import os
from datetime import datetime
from mavsdk import System
from mavsdk.offboard import OffboardError, PositionNedYaw

def load_cracks_from_excel(filename="mission2_route.xlsx"):
    cracks = []
    print(f"Reading crack coordinates from {filename} (Sheet: VisitOrder)...")
    try:
        df = pd.read_excel(filename, sheet_name="VisitOrder")
        for index, row in df.iterrows():
            if row["node_type"] != "CRACK":
                continue  
            
            cracks.append({
                "id":         str(row["crack_id"]).strip(),
                "x":          float(row["x"]),
                "y":          float(row["y"]),
                "z":          float(row["z"]),
                "crack_type": str(row["crack_type"]) if "crack_type" in df.columns and pd.notna(row.get("crack_type")) else "Unknown",
                "confidence": float(row["confidence"]) if "confidence" in df.columns and pd.notna(row.get("confidence")) else 0.0,
                "image_name": str(row["image_name"]) if "image_name" in df.columns and pd.notna(row.get("image_name")) else "",
            })
        print(f"Successfully loaded {len(cracks)} cracks in optimized order.")
        return cracks
    except Exception as e:
        print(f"ERROR reading excel file: {e}")
        return []

async def stream_and_sleep(duration, cap):
    """Waits for the drone to move while keeping the video feed live."""
    end_time = asyncio.get_event_loop().time() + duration
    latest_frame = None
    
    while asyncio.get_event_loop().time() < end_time:
        if cap and cap.isOpened():
            ret, frame = cap.read()
            if ret:
                latest_frame = frame.copy()
                cv2.imshow("Ground Station - Live Feed", frame)
        else:
            latest_frame = np.zeros((480, 640, 3), dtype=np.uint8)
            latest_frame[:] = (50, 50, 100) 
            cv2.putText(latest_frame, "WSL: WEBCAM OFFLINE", (50, 200), cv2.FONT_HERSHEY_SIMPLEX, 1.5, (0, 0, 255), 3)
            cv2.putText(latest_frame, "Using Virtual Feed", (50, 260), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 255), 2)
            cv2.imshow("Ground Station - Live Feed", latest_frame)

        cv2.waitKey(1)
        await asyncio.sleep(0.03) 
        
    return latest_frame

async def monitor_battery(drone, battery_state):
    """Background task to continuously monitor battery life."""
    async for battery in drone.telemetry.battery():
        battery_state["remaining"] = battery.remaining_percent

def generate_report(report_data, output_path="final_report.xlsx"):
    """Generates the final inspection report as an Excel file."""
    if not report_data:
        print("No inspection data to report.")
        return

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    df = pd.DataFrame(report_data)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Sheet 1: Full Report
        df.to_excel(writer, sheet_name="Report", index=False)
        ws = writer.sheets["Report"]

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(vertical="center")
                cell.border = thin_border

        col_widths = {
            "A": 12, "B": 10, "C": 10, "D": 10,
            "E": 20, "F": 12, "G": 8, "H": 8, "I": 8,
            "J": 12, "K": 35, "L": 20,
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Sheet 2: Summary
        summary_data = {
            "Metric": [
                "Total Cracks Inspected",
                "Inspection Date",
                "Crack Types Found",
                "Avg Confidence",
                "Avg Depth (mm)",
                "Min Depth (mm)",
                "Max Depth (mm)",
                "Images Saved",
            ],
            "Value": [
                len(df),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                ", ".join(sorted(df["crack_type"].unique())),
                f"{df['confidence'].mean():.4f}",
                f"{df['depth_mm'].mean():.2f}" if df["depth_mm"].sum() > 0 else "N/A (sim mode)",
                f"{df['depth_mm'].min():.2f}" if df["depth_mm"].sum() > 0 else "N/A (sim mode)",
                f"{df['depth_mm'].max():.2f}" if df["depth_mm"].sum() > 0 else "N/A (sim mode)",
                len(df[df["image_path"] != ""]),
            ],
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

        ws2 = writer.sheets["Summary"]
        for col_idx in range(1, 3):
            cell = ws2.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 40

        # Sheet 3: By crack type
        type_summary = df.groupby("crack_type").agg(
            count=("crack_type", "size"),
            avg_confidence=("confidence", "mean"),
            avg_depth_mm=("depth_mm", "mean"),
        ).reset_index()
        type_summary.columns = ["Crack Type", "Count", "Avg Confidence", "Avg Depth (mm)"]
        type_summary.to_excel(writer, sheet_name="ByType", index=False)

        ws3 = writer.sheets["ByType"]
        for col_idx in range(1, 5):
            cell = ws3.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws3.column_dimensions["A"].width = 22
        ws3.column_dimensions["B"].width = 10
        ws3.column_dimensions["C"].width = 16
        ws3.column_dimensions["D"].width = 16

    print(f"\n{'='*50}")
    print(f"  FINAL REPORT: {output_path}")
    print(f"{'='*50}")
    print(f"  Cracks inspected : {len(df)}")
    print(f"  Crack types      : {', '.join(sorted(df['crack_type'].unique()))}")
    print(f"  Avg confidence   : {df['confidence'].mean():.2f}")
    if df["depth_mm"].sum() > 0:
        print(f"  Avg depth        : {df['depth_mm'].mean():.2f} mm")
    else:
        print(f"  Depth            : N/A (sim mode — no RealSense)")
    print(f"  Report saved to  : {os.path.abspath(output_path)}")
    print(f"{'='*50}")

async def run():
    target_cracks = load_cracks_from_excel("mission2_route.xlsx")
    if not target_cracks: return

    print("SIM MODE: Starting laptop webcam for live ground feed...")
    cap = cv2.VideoCapture(0)
    await stream_and_sleep(2, cap) 

    drone = System()
    print("Connecting to Gazebo Simulator (udp://:14540)...")
    await drone.connect(system_address="udp://:14540")

    print("Waiting for drone to connect...")
    async for state in drone.core.connection_state():
        if state.is_connected:
            print("Drone discovered!")
            break

    battery_state = {"remaining": 1.0}
    asyncio.create_task(monitor_battery(drone, battery_state))

    output_folder = "img_inspection"
    os.makedirs(output_folder, exist_ok=True)

    print("-- Arming & Taking off")
    await drone.action.arm()
    await drone.action.takeoff()
    await stream_and_sleep(5, cap)

    print("-- Starting offboard mode")
    await drone.offboard.set_position_ned(PositionNedYaw(0.0, 0.0, -1.0, 0.0))
    try:
        await drone.offboard.start()
    except OffboardError as error:
        print(f"Offboard failed: {error}")
        await drone.action.disarm()
        return

    print("\n-- Starting Close-up Inspection")
    CLOSE_UP_OFFSET = 0.5 
    abort_mission = False
    report_data = []

    for crack in target_cracks:
        if abort_mission: break

        if battery_state["remaining"] < 0.20:
            print(f"CRITICAL: Battery low ({battery_state['remaining']*100:.1f}%). Aborting inspection!")
            abort_mission = True
            break

        cid, n, e, d = crack['id'], crack['x'] + CLOSE_UP_OFFSET, crack['y'], crack['z']
        print(f"\n>> Flying to Crack {cid} at [N:{n:.2f}, E:{e:.2f}, D:{d:.2f}]...")
        
        await drone.offboard.set_position_ned(PositionNedYaw(n, e, d, 0.0))
        latest_frame = await stream_and_sleep(6, cap) 
        
        print(f"Capturing webcam image of Crack {cid}...")
        image_name = f"crack_{cid}_sim_closeup.jpg"
        filepath = os.path.join(output_folder, image_name)
        
        if latest_frame is not None:
            saved_img = latest_frame.copy()
            cv2.putText(saved_img, f"CLOSE UP: Crack {cid}", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 0, 255), 2)
            cv2.imwrite(filepath, saved_img)

        report_data.append({
            "waypoint_id":  cid,
            "north":        round(n, 6),
            "east":         round(e, 6),
            "down":         round(d, 6),
            "crack_type":   crack["crack_type"],
            "confidence":   round(crack["confidence"], 4),
            "X":            0.0,
            "Y":            round(crack["y"], 6),
            "Z":            round(crack["z"], 6),
            "depth_mm":     0.0,  # no depth sensor in sim mode
            "image_path":   filepath if latest_frame is not None else "",
            "timestamp":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    # Generate final report
    generate_report(report_data, "final_report.xlsx")

    print("\n-- Inspection completed. Returning to launch...")
    await drone.offboard.set_position_ned(PositionNedYaw(0.0, 0.0, -1.0, 0.0))
    await stream_and_sleep(8, cap)

    print("-- Landing now...")
    await drone.offboard.stop()
    await drone.action.land()
    
    if cap: cap.release()
    cv2.destroyAllWindows()
    print("-- Drone landed safely.")

if __name__ == "__main__":
    asyncio.run(run())
